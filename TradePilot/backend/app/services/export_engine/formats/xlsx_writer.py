"""XLSX writer (openpyxl 기반).

특징:
    * 다중 시트 지원.
    * 컬럼 카테고리에 따른 셀 포맷(통화 / 퍼센트 / 날짜 / 정수).
    * 첫 행 고정(freeze pane) + 헤더 굵게 + 자동 너비 추정.
    * 100만 행 이상은 openpyxl 한계로 청크 단위 행 쓰기 권장.
"""
from __future__ import annotations

from io import BytesIO
from typing import Mapping

import pandas as pd

from app.services.export_engine.formats.header_map import (
    CURRENCY_COLUMNS,
    DATE_COLUMNS,
    DATETIME_COLUMNS,
    HEADER_MAP,
    NUMERIC_COLUMNS,
    PERCENT_COLUMNS,
    translate_columns,
)


# 셀 포맷 문자열 (Excel 호환)
_FORMAT_CURRENCY = '#,##0'
_FORMAT_PERCENT = '0.00%'
_FORMAT_NUMBER = '#,##0'
_FORMAT_DATE = 'yyyy-mm-dd'
_FORMAT_DATETIME = 'yyyy-mm-dd hh:mm:ss'


def _resolve_format(orig_col: str) -> str | None:
    """원본(영문) 컬럼명 기준으로 Excel number_format 결정."""
    if orig_col in CURRENCY_COLUMNS:
        return _FORMAT_CURRENCY
    if orig_col in PERCENT_COLUMNS:
        return _FORMAT_PERCENT
    if orig_col in NUMERIC_COLUMNS:
        return _FORMAT_NUMBER
    if orig_col in DATETIME_COLUMNS:
        return _FORMAT_DATETIME
    if orig_col in DATE_COLUMNS:
        return _FORMAT_DATE
    return None


def write_xlsx(sheets: Mapping[str, pd.DataFrame]) -> bytes:
    """여러 DataFrame 을 시트로 분리한 XLSX 바이트를 반환한다.

    Args:
        sheets: ``{시트명: DataFrame}`` 매핑. 시트명은 한글 가능(31자 이하).

    Returns:
        XLSX 바이너리. ``openpyxl`` 사용.
    """
    # 지연 import (openpyxl 미설치 환경에서도 import 자체는 가능하게)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # 기본 시트 제거 후 명시적으로 추가
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for raw_sheet_name, df in sheets.items():
        # 엑셀 시트 이름 제약: 최대 31자, 일부 특수문자 금지
        sheet_name = _sanitize_sheet_name(raw_sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        if df is None or df.empty:
            # 빈 데이터셋도 헤더는 남긴다 (사용자 디버깅 편의)
            translated = translate_columns(df if df is not None else pd.DataFrame())
            for col_idx, col_name in enumerate(translated.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = "A2"
            continue

        # 한글 헤더 적용
        original_cols = list(df.columns)
        korean_cols = [HEADER_MAP.get(str(c), str(c)) for c in original_cols]

        # 1행: 헤더
        for col_idx, header in enumerate(korean_cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 컬럼별 number_format 미리 계산
        col_formats: list[str | None] = [_resolve_format(c) for c in original_cols]

        # 2행 이후: 데이터
        # itertuples 가 to_records 보다 dtype 보존이 안정적
        for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_coerce(value))
                fmt = col_formats[col_idx - 1]
                if fmt is not None and value is not None:
                    cell.number_format = fmt

        # 열 너비 자동 추정 (헤더 + 첫 50행 표본)
        for col_idx, col_name in enumerate(korean_cols, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            sample_rows = df.iloc[:50, col_idx - 1] if col_idx - 1 < df.shape[1] else []
            for v in sample_rows:
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            # 한글은 1.7배 가중치, 최대 32 제한
            ws.column_dimensions[letter].width = min(max(10, int(max_len * 1.4)), 32)

        # 헤더 행 고정
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sanitize_sheet_name(name: str) -> str:
    """엑셀 시트명 제약 적용 (31자, 금지문자 치환)."""
    forbidden = set("\\/?*[]:")
    cleaned = "".join("_" if c in forbidden else c for c in name)
    return cleaned[:31] or "Sheet1"


def _coerce(value):
    """openpyxl 이 처리 못하는 타입(Decimal, datetime-naive 등) 안전 변환."""
    from datetime import date, datetime
    from decimal import Decimal

    if value is None:
        return None
    if isinstance(value, Decimal):
        # 정수면 int, 아니면 float (Excel native number)
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, (datetime, date)):
        # tz-naive 로 통일 (Excel datetime은 tz 미지원)
        if isinstance(value, datetime) and value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value
    return value
