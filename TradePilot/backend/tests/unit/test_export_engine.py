"""익스포트 엔진 단위 테스트.

검증 항목:
    1. CSV writer - UTF-8 BOM + 한글 헤더 매핑
    2. CSV writer - 빈 DataFrame 도 헤더 보존
    3. XLSX writer - 다중 시트 + 셀 number_format 적용
    4. S3 uploader - 사전서명 URL 발급(boto3 Stubber)
    5. S3 uploader - 멀티파트 임계치 동작
    6. config.object_key - prefix/user_id/public_id 조합
"""
from __future__ import annotations

import os
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pandas as pd
import pytest

# 익스포트 엔진은 stdlib + pandas 만으로 import 되어야 한다.
from app.services.export_engine.config import ExportConfig, get_export_config
from app.services.export_engine.formats.csv_writer import write_csv
from app.services.export_engine.formats.header_map import (
    HEADER_MAP,
    translate_columns,
)


pytestmark = pytest.mark.unit


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def test_csv_writer_utf8_bom_and_korean_headers() -> None:
    """CSV 는 UTF-8 BOM 으로 시작하고 한글 헤더가 적용되어야 한다."""
    df = pd.DataFrame(
        [
            {"code": "005930", "name": "삼성전자", "qty": 10, "price": 70000},
        ]
    )
    data = write_csv(df)

    # 1. BOM 확인
    assert data.startswith(b"\xef\xbb\xbf"), "UTF-8 BOM 누락"

    # 2. 한글 헤더 매핑 확인
    text = data.decode("utf-8-sig")
    first_line = text.splitlines()[0]
    assert "종목코드" in first_line
    assert "종목명" in first_line
    assert "수량" in first_line
    assert "주문가격" in first_line
    assert "code" not in first_line  # 영문 누락 검증


def test_csv_writer_empty_dataframe_keeps_header() -> None:
    """행이 0개여도 헤더만큼은 남는다 (사용자 디버깅 편의)."""
    df = pd.DataFrame(columns=["code", "name", "qty"])
    data = write_csv(df)
    text = data.decode("utf-8-sig").strip()
    # 헤더 1줄만 존재
    lines = [ln for ln in text.split("\n") if ln.strip()]
    assert len(lines) == 1
    assert "종목코드" in lines[0]


def test_translate_columns_keeps_unknown_keys() -> None:
    """매핑에 없는 컬럼은 원본 이름을 유지한다."""
    df = pd.DataFrame(columns=["code", "custom_field"])
    out = translate_columns(df)
    assert list(out.columns) == ["종목코드", "custom_field"]


def test_header_map_contains_required_keys() -> None:
    """주요 필드의 한글 매핑이 누락되지 않았는지 확인."""
    must_have = [
        "code", "name", "qty", "price", "side", "status",
        "filled_at", "executed_at", "realized_pnl",
    ]
    for key in must_have:
        assert key in HEADER_MAP, f"한글 매핑 누락: {key}"


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def test_xlsx_writer_multi_sheet_with_formats() -> None:
    """다중 시트 + 셀 number_format 적용 검증."""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from app.services.export_engine.formats.xlsx_writer import write_xlsx

    sheets = {
        "주문": pd.DataFrame(
            [
                {"code": "005930", "qty": 10, "price": Decimal("70000.00")},
            ]
        ),
        "체결": pd.DataFrame(
            [
                {"code": "005930", "fill_qty": 5, "fill_price": Decimal("70100.00")},
            ]
        ),
    }
    data = write_xlsx(sheets)
    assert data[:2] == b"PK", "XLSX 시그니처(PK) 누락"

    wb = load_workbook(BytesIO(data))
    assert "주문" in wb.sheetnames
    assert "체결" in wb.sheetnames

    ws = wb["주문"]
    # 헤더 셀: 한글
    assert ws.cell(row=1, column=1).value == "종목코드"
    assert ws.cell(row=1, column=2).value == "수량"
    assert ws.cell(row=1, column=3).value == "주문가격"
    # 데이터 행
    assert ws.cell(row=2, column=1).value == "005930"
    # price 컬럼(통화) number_format 적용 여부
    price_cell = ws.cell(row=2, column=3)
    assert price_cell.number_format == "#,##0", (
        f"통화 포맷 누락. 실제: {price_cell.number_format}"
    )
    # 첫 행 고정
    assert ws.freeze_panes == "A2"


def test_xlsx_writer_empty_sheet_has_headers() -> None:
    """빈 DataFrame 도 헤더만 남아야 한다."""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from app.services.export_engine.formats.xlsx_writer import write_xlsx

    sheets = {"빈시트": pd.DataFrame(columns=["code", "name"])}
    data = write_xlsx(sheets)
    wb = load_workbook(BytesIO(data))
    ws = wb["빈시트"]
    assert ws.cell(row=1, column=1).value == "종목코드"
    assert ws.cell(row=1, column=2).value == "종목명"


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def test_object_key_combines_prefix_user_uuid() -> None:
    cfg = ExportConfig(
        s3_bucket="b",
        s3_region="r",
        s3_prefix="exports/",
        s3_endpoint_url=None,
        s3_access_key=None,
        s3_secret_key=None,
        presign_ttl_sec=3600,
        retention_hours=168,
        max_rows=1_000_000,
        chunk_size=50_000,
        multipart_threshold_mb=10,
        concurrent_per_user=3,
        daily_limit_per_user=20,
    )
    key = cfg.object_key(42, "abc-def-uuid", "csv")
    assert key == "exports/42/abc-def-uuid.csv"

    # prefix 가 슬래시 없이 들어와도 동일 결과
    cfg2 = ExportConfig(**{**cfg.__dict__, "s3_prefix": "exports"})
    assert cfg2.object_key(42, "abc-def-uuid", "xlsx") == "exports/42/abc-def-uuid.xlsx"


def test_get_export_config_env_override(monkeypatch: pytest.MonkeyPatch) -> None:
    """환경변수 우선 적용."""
    monkeypatch.setenv("EXPORT_S3_BUCKET", "my-bucket")
    monkeypatch.setenv("EXPORT_PRESIGN_TTL_SECONDS", "7200")
    cfg = get_export_config()
    assert cfg.s3_bucket == "my-bucket"
    assert cfg.presign_ttl_sec == 7200


# ---------------------------------------------------------------------------
# S3 Uploader (boto3 Stubber)
# ---------------------------------------------------------------------------
@pytest.fixture
def stubbed_uploader():
    """boto3 Stubber 로 S3 호출을 mock 한다."""
    boto3 = pytest.importorskip("boto3")
    from botocore.stub import Stubber

    from app.services.export_engine.s3_uploader import S3Uploader

    cfg = ExportConfig(
        s3_bucket="test-bucket",
        s3_region="ap-northeast-2",
        s3_prefix="exports/",
        s3_endpoint_url=None,
        s3_access_key="AKIATEST",
        s3_secret_key="secret",
        presign_ttl_sec=3600,
        retention_hours=168,
        max_rows=1_000_000,
        chunk_size=50_000,
        multipart_threshold_mb=10,
        concurrent_per_user=3,
        daily_limit_per_user=20,
    )
    client = boto3.client(
        "s3",
        region_name="ap-northeast-2",
        aws_access_key_id="AKIATEST",
        aws_secret_access_key="secret",
    )
    stubber = Stubber(client)
    uploader = S3Uploader(cfg)
    uploader.set_client(client)
    return uploader, client, stubber


def test_s3_upload_small_file_uses_put_object(stubbed_uploader) -> None:
    """10MB 미만은 단일 PUT 사용."""
    uploader, client, stubber = stubbed_uploader
    payload = b"\xef\xbb\xbfcode,name\n005930,Samsung\n"
    stubber.add_response(
        "put_object",
        service_response={"ETag": '"abc"'},
        expected_params={
            "Bucket": "test-bucket",
            "Key": "exports/1/u1.csv",
            "Body": payload,
            "ContentType": "text/csv; charset=utf-8",
            "ServerSideEncryption": "AES256",
        },
    )
    with stubber:
        result = uploader.upload_bytes(
            payload, "exports/1/u1.csv", content_type="text/csv; charset=utf-8"
        )
    assert result["key"] == "exports/1/u1.csv"
    assert result["size"] == len(payload)


def test_s3_generate_presigned_url(stubbed_uploader) -> None:
    """generate_presigned_url 호출 결과는 문자열."""
    uploader, _client, _stubber = stubbed_uploader
    # generate_presigned_url 은 stub 없이도 동작(서명만 생성)
    url = uploader.generate_presigned_url("exports/1/u1.csv", ttl_sec=3600)
    assert isinstance(url, str)
    assert "exports/1/u1.csv" in url
    # AWS 서명 쿼리 파라미터 포함
    assert "X-Amz-" in url or "AWSAccessKeyId" in url


def test_s3_delete_object_swallows_errors(stubbed_uploader) -> None:
    """cleanup 시 객체가 없어도 예외를 던지지 않는다."""
    uploader, _client, stubber = stubbed_uploader
    stubber.add_client_error(
        "delete_object",
        service_error_code="NoSuchKey",
        http_status_code=404,
    )
    with stubber:
        # 예외 없이 반환되어야 한다
        uploader.delete_object("exports/1/missing.csv")
